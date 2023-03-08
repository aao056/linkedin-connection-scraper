from argParserSetup import args
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.by import By
import requests
import os
from dotenv import load_dotenv
from pathlib import Path

load_dotenv()
EMAIL = os.getenv('EMAIL')
PASSWORD = os.getenv('PASSWORD')

# Set the filename and path for the file
filename = "data.xlsx"
file_path = Path.cwd() / filename

# Check if the file already exists
if os.path.exists(file_path):
    # If the file exists, load it into openpyxl
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook['Sheet']
    workbook.save(file_path)
else:
    # If the file doesn't exist, create a new workbook and save it to the file
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    workbook.save(file_path)

worksheet.cell(row=1, column=1, value='First Name')
worksheet.cell(row=1, column=2, value='Last Name')
worksheet.cell(row=1, column=3, value='Gender')
worksheet.cell(row=1, column=4, value='Location')
worksheet.cell(row=1, column=5, value='LinkedIn URL')

driver = webdriver.Chrome(service=Service('/usr/local/bin/chromedriver'))
driver.get("https://www.linkedin.com/")
time.sleep(4)
email_input = driver.find_element(By.XPATH, "//input[@autocomplete='username']")
email_input.send_keys(str(EMAIL))
password_input = driver.find_element(By.XPATH, "//input[@autocomplete='current-password']")
password_input.send_keys(str(PASSWORD))
submit_button = driver.find_element(By.XPATH, "//button[@type='submit']")
submit_button.submit()
driver.get(args.target_url)
time.sleep(5)
connections_link = driver.find_element(By.XPATH, "//li/a[contains(@href, 'connection')]").get_attribute('href')
driver.get(connections_link)

for i in range(args.start, args.end):
    url = connections_link + f'&page={i}'
    driver.get(url)
    time.sleep(3)
    people = driver.find_elements(By.XPATH, "//li[contains(@class, 'result')]")
    offset = i * 10 - 9
    for row, item in enumerate(people):
        while True:
            try:
                name = item.find_element(By.XPATH, './/span[@dir="ltr"]/span[1]')
                first_name = name.text.split(" ")[0]
                last_name = name.text.split(" ")[1]
                GENDERIZE_URL = f'https://api.genderize.io?name={name.text.split(" ")[0]}'
                GENDERIZE_RESPONSE = requests.get(GENDERIZE_URL)
                data = GENDERIZE_RESPONSE.json()
                gender = data['gender']
                location = item.find_element(By.XPATH, './/div[contains(@class, "secondary-subtitle")]')
                link = item.find_element(By.XPATH, './/a[@class="app-aware-link "][span]').get_attribute('href').split('?')[0]
                worksheet.cell(row=row + 1 + offset, column=1, value=first_name)
                worksheet.cell(row=row + 1 + offset, column=2, value=last_name)
                worksheet.cell(row=row + 1 + offset, column=3, value=gender)
                worksheet.cell(row=row + 1 + offset, column=4, value=location.text)
                worksheet.cell(row=row + 1 + offset, column=5, value=link)
            except BaseException as error:
                print('An exception occurred: {}'.format(error))
                continue
            break

workbook.save(file_path)
driver.close()